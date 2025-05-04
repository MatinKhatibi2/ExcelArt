'''
Excel Image Art Generator

Converts images to Excel color grids with configurable resolution using pixel averaging.
'''

import openpyxl as xls
from PIL import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_STEP = 5
CELL_WIDTH = 3
ROW_HEIGHT = 20


def process(path: str, step: int = DEFAULT_STEP) -> list[list[tuple[int, int, int]]]:
    '''
    Process image into color blocks

    Args:
        path: Path to source image
        step: Pixel grouping size (controls output resolution)

    Returns:
        2D list of RGB color tuples
    '''

    image = Image.open(path)
    w, h = image.size
    colors = []

    for y in range(0, h, step):
        row_colors = []
        for x in range(0, w, step):
            # Calculate block boundaries
            box = (
                x,
                y,
                min(x+step, w),
                min(y+step, h)
            )

            block = image.crop(box)

            row_colors.append(calculate_average_color(block))
        colors.append(row_colors)

    return colors


def calculate_average_color(block: Image.Image) -> tuple[int, int, int]:
    '''Calculate average color of an image block'''
    r_total = g_total = b_total = 0
    pixel_count = block.width * block.height
    for x in range(block.width):
        for y in range(block.height):
            r, g, b = block.getpixel((x, y))[:3]
            r_total += r
            g_total += g
            b_total += b
    return (r_total // pixel_count, g_total // pixel_count, b_total // pixel_count)


def create_excel(colors: list[list[tuple[int, int, int]]], result_path: str):
    """
    Generate Excel file from color grid

    Args:
        colors: 2D list of RGB tuples from process()
        output_path: Path to save Excel file
    """

    wb = xls.Workbook()
    ws = wb.active

    for i, row in enumerate(colors, 1):
        for j, color in enumerate(row, 1):

            ws.column_dimensions[get_column_letter(j)].width = CELL_WIDTH
            ws.row_dimensions[i].height = ROW_HEIGHT

            hex_color = f"{color[0]:02X}{color[1]:02X}{color[2]:02X}"

            ws.cell(row=i, column=j).fill = PatternFill(
                start_color=hex_color,
                end_color=hex_color,
                fill_type="solid"
            )
    wb.save(result_path)


if __name__ == "__main__":
    # Example usage
    color_data = process(
        r"C:\Users\aliot\OneDrive\Desktop\test-matin\Maktab\Excel\Excel-Image-Art\examples\input\img.jpg", step=20)
    create_excel(
        color_data, r"C:\Users\aliot\OneDrive\Desktop\test-matin\Maktab\Excel\Excel-Image-Art\examples\output\res.xlsx")
